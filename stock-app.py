import os
import time
import math
import datetime
import openpyxl
import pandas as pd
import numpy as np
import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import requests as rq
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
start_time = time.time()

monthdata = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12
}


class GetDynamic:

    @st.cache(show_spinner=False)
    def get_ticker(self, static_url, stock_name):

        formatted_stock_name = stock_name.replace(' ', '').upper()

        loc = os.getcwd()
        file_loc = os.path.join(loc, 'cache.xlsx')

        if os.path.isfile(file_loc):
            wb = load_workbook('cache.xlsx')

        else:
            wb = Workbook()

        sheet = wb.active
        row_size = sheet.max_row

        for row in range(1, row_size+1):

            if sheet['A'+str(row)].value == formatted_stock_name:
                ticker_symbol = sheet['B'+str(row)].value
                break

            else:
                ticker_symbol = None

        if ticker_symbol != None:
            return ticker_symbol

        else:

            user_agent = "C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
            options = webdriver.ChromeOptions()
            options.headless = True
            options.add_argument(f'user-agent={user_agent}')
            options.add_argument('--disable-gpu')

            driver = webdriver.Chrome(
                executable_path="chromedriver.exe", options=options)

            driver.get(static_url)

            try:
                seacrh_box = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, 'yfin-usr-qry'))
                )
            except:
                driver.quit()

            seacrh_box.send_keys(stock_name)
            search = driver.find_element_by_id('search-button')
            search.click()
            hist_data1 = None
            try:
                hist_data1 = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located(
                        (By.ID, 'quote-nav'))
                )
            except:
                driver.quit()

            if hist_data1 == None:
                dynamic_url = None
            else:
                hist_data2 = hist_data1.find_element_by_tag_name(
                    'ul').find_elements_by_tag_name('li')
                hist_data2[4].click()
                time.sleep(1)
                dynamic_url = driver.current_url

            driver.quit()

            headers = {
                'User-Agent': 'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'}
            html = rq.get(dynamic_url, headers=headers).text

            soup = BeautifulSoup(html, 'lxml')

            name = soup.find(
                'h1', class_='D(ib) Fz(18px)').text.replace(' ', '')
            name_split = []
            for letter in name:
                name_split.append(letter)

            for i in range(len(name_split)):
                if name_split[i] == '(':
                    start_index = i
                elif name_split[i] == ')':
                    end_index = i
            ticker = name_split[start_index+1:end_index]
            ticker_symbol = ""
            ticker_symbol = ticker_symbol.join(ticker)
            sheet.append((formatted_stock_name, ticker_symbol))
            wb.save('cache.xlsx')
            return ticker_symbol

    @st.cache(show_spinner=False)
    def get_dynamic(self, ticker):
        dynamic = f'https://in.finance.yahoo.com/quote/{ticker}/history?p={ticker}'
        return dynamic


class Regression:
    class LogisticRegression:

        def __init__(self, lr=0.001, iters=1000):
            self.lr = lr
            self.n_iters = iters
            self.weights = None
            self.bias = None

        def sigmoid(self, x):
            return 1 / (1 + np.exp(-x))

        def train(self, xtrain, ytrain):
            n_samples, n_features = xtrain.shape
            self.weights = np.zeros(n_features)
            self.bias = 0

            for _ in range(self.n_iters):
                linear = np.dot(xtrain, self.weights) + self.bias
                y_predicted = self.sigmoid(linear)

                dw = (1 / n_samples) * np.dot(xtrain.T, (y_predicted - ytrain))
                db = (1 / n_samples) * np.sum(y_predicted - ytrain)

                self.weights -= self.lr * dw
                self.bias -= self.lr * db

        def predict(self, X):
            linear = np.dot(X, self.weights) + self.bias
            y_predicted = self.sigmoid(linear)
            y_predicted_cls = [1 if i > 0.5 else 0 for i in y_predicted]
            y_real = [-i if i < 0.5 else i for i in y_predicted]
            return np.array(y_predicted_cls), np.array(y_real)

        def get_log_data(self, data, Close):

            df = data['Close'].pct_change()*100
            df = df.reset_index()
            for i in range(1, 6):
                df['Lag'+str(i)] = df['Close'].shift(i)
            df = df.dropna(axis=0)
            df['Volume'] = data['Volume'][6:].values/1000000000
            df['Direction'] = [1 if i > 0 else 0 for i in df['Close']]
            df = df.reset_index()

            X = df[['Close', 'Lag1', 'Lag2', 'Lag3', 'Lag4', 'Lag5', 'Volume']]
            y = df['Direction']

            cls = data[['Date', 'Close', 'Volume']]
            cls['Pos'] = [i for i in range(0, len(cls))]

            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.2, random_state=1234)

            self.train(X_train, y_train)
            pred, dump = self.predict(X_test)
            accur = (np.sum(pred == y_test))/len(y_test)*100
            print(f"Accuracy: {accur}%")

            val = cls['Pos'][cls['Close'] == Close]
            prev_close = [float(cls['Close'][val+i]) for i in range(0, 7)]
            prev_close = pd.DataFrame(prev_close)

            lag = np.array(pd.DataFrame.pct_change(prev_close)*100).reshape(7,)
            lags = [i for i in lag if math.isnan(i) == False]
            volume = float(cls['Volume'][val]/1000000000)
            print(lags)
            fig, real = self.predict(
                [[lags[0], lags[1], lags[2], lags[3], lags[4], lags[5], volume]])

            return fig[0], real[0], accur

    class LinearRegression:

        def __init__(self, lr=0.001, iters=1000):
            self.lr = lr
            self.iters = iters
            self.weights = None
            self.bias = None

        def train(self, xtrain, ytrain):
            n_samples, n_features = np.shape(xtrain)
            self.weights = np.zeros(n_features)
            self.bias = 0

            for _ in range(self.iters):
                y_cap = np.dot(xtrain, self.weights)+self.bias

                dw = (1/n_samples)*(np.dot(xtrain.T, (y_cap-ytrain)))
                db = (1/n_samples)*(np.sum(y_cap-ytrain))

                self.weights -= self.lr*dw
                self.bias -= self.lr*db

        def predict(self, xtest):
            y_cap = np.dot(xtest, self.weights)+self.bias
            return y_cap

        def mse(self, pred, ytest):
            error = np.mean((ytest-pred)**2)
            return error

        def get_linear_data(self, data, Open):
            logistic = regression.LogisticRegression(
                lr=0.01, iters=1000)
            linear = regression.LinearRegression(lr=0.00001, iters=1000)

            data['Direction'] = [1 if data['Close'][i] >
                                 data['Open'][i] else 0 for i in range(len(data))]

            X = data[['Open', 'Direction']]
            y = data['Close']

            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.2, random_state=1234)

            linear.train(X_train, y_train)
            pred = linear.predict(X_test)
            error = linear.mse(y_test, pred)
            print(f"Error : {error} ")

            dir, val, acc = logistic.get_log_data(data, Open)
            val *= 50
            print(f"{dir}::{val}")
            return linear.predict([[Open, val]])[0], error, acc


class Data:
    def get_hist_data(self, name, x, y, method='r'):

        sy = x[0]
        sm = x[1]
        sd = x[2]

        ey = y[0]
        em = y[1]
        ed = y[2]

        start_date = int(time.mktime(
            datetime.datetime(sy, sm, sd, 23, 59).timetuple()))
        end_date = int(time.mktime(
            datetime.datetime(ey, em, ed, 23, 59).timetuple()))
        interval = '1d'
        query = f'https://query1.finance.yahoo.com/v7/finance/download/{name}?period1={start_date}&period2={end_date}&interval={interval}&events=history&includeAdjustedClose=true'
        headers = {
            'User-Agent': 'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'}
        html = rq.get(query, headers=headers).text

        df = pd.read_csv(query)
        df = df.dropna(axis=0)
        if method == 'p':
            df = df.iloc[::-1]
        else:
            pass
        df = df.reset_index()
        df = df[['Date', 'Open', 'High', 'Low', 'Close', 'Volume']]

        return df

    def get_name(self, ticker):
        url = f"https://finance.yahoo.com/quote/{ticker}/history?p={ticker}"
        headers = {
            'User-Agent': 'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'}

        html = rq.get(url, headers=headers).text

        soup = BeautifulSoup(html, 'lxml')

        sname = soup.find('h1', class_="D(ib) Fz(18px)").text
        den = soup.find(
            'div', class_="C($tertiaryColor) Fz(12px)").find('span').text
        return sname, den

    def realtime_learn(self, ticker, et):

        url = f"https://in.finance.yahoo.com/quote/{ticker}/history?p={ticker}"
        headers = {
            'User-Agent': 'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe'}

        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'Time'
        sheet['B1'] = 'Price'

        html = rq.get(url, headers=headers).text
        soup = BeautifulSoup(html, 'lxml')

        now = time.strftime("%H:%M")
        while now != et:
            time.sleep(2)
            price = float(soup.find(
                'span', class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").text)
            ct = time.strftime("%H:%M:%S")
            val = (ct, price)
            sheet.append(val)
            now = time.strftime("%H:%M")

        filename = ticker + '.xlsx'
        wb.save(filename)
        df = pd.read_excel(filename)
        return df


class Tools(GetDynamic, Data, Regression.LinearRegression, Regression.LogisticRegression):

    def show_data(self, ticker):
        sname, den = self.get_name(ticker)
        monthlist = ("", "January", "February", "March", "April",
                     "May", "June", "July", "August", "September", "October", "November", "December")

        years = range(2021, 1999, -1)
        sidebar_years = [""]
        for year in years:
            sidebar_years.append(year)

        sy = st.sidebar.selectbox("Start Year", sidebar_years)
        try:
            sm = monthdata[st.sidebar.selectbox("Start Month", monthlist)]
        except KeyError:
            st.write("Provide start and end dates!")

        ey = st.sidebar.selectbox("End Year", sidebar_years)
        try:
            em = monthdata[st.sidebar.selectbox("End Month", monthlist)]
        except KeyError:
            st.write("")

        try:
            start = [sy, sm, 1]
            end = [ey, em, 1]
        except NameError:
            st.write("")
        try:
            if sm == em and sy >= ey:
                st.write("### **Invalid date**, _start and end dates are same!_")
            elif sy > ey:
                st.write(
                    "### **Invalid date**, _end date must be after start date!_ ")
            else:
                try:
                    df = self.get_hist_data(ticker, start, end)
                    st.write(f"## Showing data for _**{sname}**_")
                    st.write(den)
                    st.write("# Data")
                    st.dataframe(df)
                    return df
                except NameError:
                    st.write("")
                    return None
        except:
            st.write("")

    def show_comparison(self, df1, df2):
        df1['New'] = df2['Close']
        try:
            st.write("""# Visualization
                """)
            st.line_chart(df1[['Close', 'New']], use_container_width=True)
        except TypeError:
            st.write("")

    def show_graph(self, df, method='norm'):
        if method == 'norm':
            c = 'Close'
        elif method == 'rltp':
            c = 'Price'
        try:
            st.write("""# Visualization
                """)
            st.line_chart(df[c], use_container_width=True)
        except TypeError:
            st.write("")

    def show_prediction(self, ticks):
        monthlist = ("January", "February", "March", "April",
                     "May", "June", "July", "August", "September", "October", "November", "December")

        years = range(2021, 2019, -1)
        days = range(1, 32)
        sm = 7
        sd = 15
        try:
            sy = st.sidebar.selectbox("Year", years)
            sm = monthdata[st.sidebar.selectbox("Month", monthlist)]
        except KeyError:
            st.write("")

        try:
            sd = st.sidebar.selectbox("Day", days)
        except KeyError:
            st.write("Provide date for prediction!")
        start = [sy, sm, sd]
        end = [sy-2, sm, sd]
        data = self.get_hist_data(ticks, end, start, 'p')
        Close = float(data['Close'][0])
        try:
            predicted_close, error, acc = self.get_linear_data(
                data, Close)
            error = '%.2f' % error
        except:
            st.write("### Try using a different date!")

        st.sidebar.write(f"### Accuracy : {acc}%")
        st.sidebar.write(f"### Error : {error}")
        Close = '%0.2f' % Close
        predicted_close = '%0.2f' % predicted_close
        sname, den = self.get_name(ticks)
        st.write(f"## Showing data for _**{sname}**_")
        st.write(den)

        st.write(f"""## Previous Close
        {Close}""")

        st.write(f"""## Predicted Close for {sd}/{sm}/{sy}
        {predicted_close}""")

    def show_compare_info(self, ticker1, ticker2):
        monthlist = ("", "January", "February", "March", "April",
                     "May", "June", "July", "August", "September", "October", "November", "December")

        years = range(2021, 1999, -1)
        sidebar_years = [""]
        for year in years:
            sidebar_years.append(year)

        sy = st.sidebar.selectbox("Start Year", sidebar_years)
        try:
            sm = monthdata[st.sidebar.selectbox("Start Month", monthlist)]
        except KeyError:
            st.write("Provide start and end dates!")

        ey = st.sidebar.selectbox("End Year", sidebar_years)
        try:
            em = monthdata[st.sidebar.selectbox("End Month", monthlist)]
        except KeyError:
            st.write("")

        try:
            start = [sy, sm, 1]
            end = [ey, em, 1]
        except NameError:
            st.write("")
        try:
            if sm == em and sy >= ey:
                st.write(
                    "### **Invalid date**, _start and end dates are same!_")
            elif sy > ey:
                st.write(
                    "### **Invalid date**, _end date must be after start date!_ ")
            else:
                try:
                    df1 = data.get_hist_data(ticker1, start, end)
                    df2 = data.get_hist_data(ticker2, start, end)
                except NameError:
                    st.write("")
                    return None
                nm1, den1 = data.get_name(ticker1)
                nm2, den2 = data.get_name(ticker2)
                self.show_comparison(df1, df2)
                st.write(f"Blue   : {ticker1}")
                st.write(f"Orange : {ticker2}")

                var1 = -((df1['Close'][0]-df1['Close']
                          [len(df1)-1])/df1['Close'][0])*100
                var2 = -((df2['Close'][0]-df2['Close']
                          [len(df2)-1])/df2['Close'][0])*100
                d1 = ['Decrease' if var1 < 0 else 'Increase']
                d2 = ['Decrease' if var2 < 0 else 'Increase']
                var1 = ['%.2f' % var1 if var1 > 0 else '%.2f' % -var1]
                var2 = ['%.2f' % var2 if var2 > 0 else '%.2f' % -var2]
                st.write(f"## _**{nm1}**_")
                st.write(den1)
                st.dataframe(df1)
                st.write(f"""### Growth
                {var1[0]}% {d1[0]}""")

                st.write(f"## _**{nm2}**_")
                st.write(den2)
                st.dataframe(df2)
                st.write(f"""### Growth
                {var2[0]}% {d2[0]}""")

        except:
            st.write("")

    def show_realtime_learn(self, ticker):
        df = None
        st.write("Input time to stop learning")
        try:
            hour = st.text_input("Hour: ")
        except:
            st.write("")
        minute = st.text_input("Minute: ")
        et = hour+':'+minute
        button = st.button("Click to start")

        if button == True:
            st.write(f""" ## Learning data till {et}
# Please do not close this window keep it _**minimized!**_""")
            df = data.realtime_learn(ticker, et)

        st.dataframe(df)
        try:
            var = ((df['Price'][0]-df['Price'][len(df)-1])/df['Price'][0])*100
            st.write(f"""  # Total Variance
                    {var} %""")
        except:
            st.write("")

        return df


static = 'https://in.finance.yahoo.com/'
regression = Regression()
fetch = GetDynamic()
data = Data()
tools = Tools()


operation = st.sidebar.selectbox(
    "Choose Operation", (" ", "Predict Close Price", "Realtime learning", "Compare Stocks", "Get Historical data"))

if operation == " ":
    st.write("""# _Welcome!_


# _Choose an Operation you need to perform in the sidebar and Input the stock name!_
 """)
else:
    ticker = st.text_input("Ticker", "TSLA")
    dynamic = fetch.get_dynamic(ticker)

    if operation == "Realtime learning":
        df = tools.show_realtime_learn(ticker)
        tools.show_graph(df, 'rltp')

    elif operation == "Get Historical data":
        df = tools.show_data(ticker)
        tools.show_graph(df)

    elif operation == "Predict Close Price":
        df = tools.show_prediction(ticker)

    elif operation == "Compare Stocks":
        ticker1 = ticker
        ticker2 = st.text_input("Stock to compare with", "MSFT")
        tools.show_compare_info(ticker1, ticker2)

passed_time = time.time() - start_time
print(f"Request completed in : {passed_time} seconds")
